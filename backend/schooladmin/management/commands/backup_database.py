"""
Django management command to backup the database
Exports all tables to Excel and creates a ZIP archive

Usage:
    python manage.py backup_database
    python manage.py backup_database --email user@example.com
"""

from django.core.management.base import BaseCommand
from django.conf import settings
from django.apps import apps
from django.core.mail import EmailMessage
from datetime import datetime
import os
import shutil
import zipfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from django.utils import timezone as django_timezone


class Command(BaseCommand):
    help = 'Backup database to Excel files and create ZIP archive'

    def add_arguments(self, parser):
        parser.add_argument(
            '--email',
            type=str,
            help='Email address to send backup to (optional)',
        )
        parser.add_argument(
            '--no-zip',
            action='store_true',
            help='Skip creating ZIP file',
        )

    def handle(self, *args, **options):
        self.stdout.write(self.style.SUCCESS('=' * 70))
        self.stdout.write(self.style.SUCCESS('DATABASE BACKUP STARTED'))
        self.stdout.write(self.style.SUCCESS('=' * 70))

        # Create timestamp for backup folder
        timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')

        # Create backup directory structure
        base_backup_dir = os.path.join(settings.BASE_DIR.parent, 'backups')
        backup_dir = os.path.join(base_backup_dir, f'backup_{timestamp}')

        os.makedirs(backup_dir, exist_ok=True)
        self.stdout.write(f'Created backup directory: {backup_dir}')

        # 1. Export all tables to Excel
        excel_file = self.export_to_excel(backup_dir, timestamp)

        # 2. Copy raw SQLite database
        db_backup = self.backup_sqlite_db(backup_dir)

        # 3. Create backup summary
        summary_file = self.create_summary(backup_dir, timestamp)

        # 4. Create ZIP archive
        zip_file = None
        if not options['no_zip']:
            zip_file = self.create_zip_archive(backup_dir, timestamp)

        # 5. Send email if requested
        if options['email']:
            self.send_backup_email(options['email'], zip_file or excel_file, timestamp)

        # Print summary
        self.stdout.write(self.style.SUCCESS('\n' + '=' * 70))
        self.stdout.write(self.style.SUCCESS('BACKUP COMPLETED SUCCESSFULLY'))
        self.stdout.write(self.style.SUCCESS('=' * 70))
        self.stdout.write(f'\nBackup location: {backup_dir}')
        if excel_file:
            self.stdout.write(f'Excel file: {os.path.basename(excel_file)}')
        if db_backup:
            self.stdout.write(f'SQLite backup: {os.path.basename(db_backup)}')
        if zip_file:
            self.stdout.write(f'ZIP archive: {os.path.basename(zip_file)}')
        if summary_file:
            self.stdout.write(f'Summary: {os.path.basename(summary_file)}')
        self.stdout.write(self.style.SUCCESS('\n' + '=' * 70 + '\n'))

    def export_to_excel(self, backup_dir, timestamp):
        """Export all database tables to a single Excel file with multiple sheets"""
        self.stdout.write('\n[1/4] Exporting database tables to Excel...')

        excel_path = os.path.join(backup_dir, f'database_backup_{timestamp}.xlsx')
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Get all models from all apps
        models_to_export = self.get_models_to_export()

        total_tables = len(models_to_export)
        total_records = 0

        for idx, (app_label, model) in enumerate(models_to_export, 1):
            model_name = model._meta.model_name
            sheet_name = f"{app_label}_{model_name}"[:31]  # Excel sheet name limit

            self.stdout.write(f'  [{idx}/{total_tables}] Exporting {app_label}.{model_name}...', ending='')

            # Create worksheet
            ws = wb.create_sheet(title=sheet_name)

            # Get all field names
            fields = [field.name for field in model._meta.fields]

            # Write header row
            for col_idx, field_name in enumerate(fields, 1):
                cell = ws.cell(row=1, column=col_idx, value=field_name)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

            # Write data rows
            queryset = model.objects.all()
            record_count = 0

            for row_idx, obj in enumerate(queryset, 2):
                for col_idx, field_name in enumerate(fields, 1):
                    try:
                        value = getattr(obj, field_name)
                        # Handle special types
                        if hasattr(value, 'id'):  # Foreign key
                            value = str(value)
                        elif isinstance(value, (list, dict)):  # JSON fields
                            value = str(value)
                        elif isinstance(value, datetime):  # Datetime with timezone
                            # Remove timezone info for Excel compatibility
                            if value.tzinfo is not None:
                                value = value.replace(tzinfo=None)
                        elif value is None:
                            value = ''
                        ws.cell(row=row_idx, column=col_idx, value=value)
                    except Exception as e:
                        ws.cell(row=row_idx, column=col_idx, value=f'ERROR: {str(e)}')

                record_count += 1

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            total_records += record_count
            self.stdout.write(self.style.SUCCESS(f' {record_count} records'))

        # Save workbook
        wb.save(excel_path)

        self.stdout.write(self.style.SUCCESS(f'\n  ✓ Excel file created: {total_tables} tables, {total_records} total records\n'))
        return excel_path

    def backup_sqlite_db(self, backup_dir):
        """Copy the raw SQLite database file"""
        self.stdout.write('[2/4] Backing up SQLite database file...')

        db_path = settings.DATABASES['default']['NAME']

        if not os.path.exists(db_path):
            self.stdout.write(self.style.WARNING('  ⚠ SQLite database file not found'))
            return None

        db_backup_path = os.path.join(backup_dir, 'db.sqlite3')
        shutil.copy2(db_path, db_backup_path)

        file_size = os.path.getsize(db_backup_path) / (1024 * 1024)  # Convert to MB
        self.stdout.write(self.style.SUCCESS(f'  ✓ Database copied: {file_size:.2f} MB\n'))

        return db_backup_path

    def create_summary(self, backup_dir, timestamp):
        """Create a text summary of the backup"""
        self.stdout.write('[3/4] Creating backup summary...')

        summary_path = os.path.join(backup_dir, 'backup_summary.txt')

        models_to_export = self.get_models_to_export()

        with open(summary_path, 'w') as f:
            f.write('=' * 70 + '\n')
            f.write('SCHOOL MANAGEMENT SYSTEM - DATABASE BACKUP SUMMARY\n')
            f.write('=' * 70 + '\n\n')
            f.write(f'Backup Date: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\n')
            f.write(f'Backup ID: {timestamp}\n\n')

            f.write('TABLES BACKED UP:\n')
            f.write('-' * 70 + '\n')

            total_records = 0
            for app_label, model in models_to_export:
                count = model.objects.count()
                total_records += count
                f.write(f'{app_label}.{model._meta.model_name:<30} {count:>10} records\n')

            f.write('-' * 70 + '\n')
            f.write(f'{"TOTAL":<30} {total_records:>10} records\n')
            f.write('=' * 70 + '\n\n')

            f.write('FILES INCLUDED:\n')
            f.write('  - database_backup_{}.xlsx (All tables in Excel format)\n'.format(timestamp))
            f.write('  - db.sqlite3 (Raw SQLite database)\n')
            f.write('  - backup_summary.txt (This file)\n\n')

            f.write('RESTORATION:\n')
            f.write('  To restore from SQLite backup:\n')
            f.write('    1. Stop the Django server\n')
            f.write('    2. Replace backend/db.sqlite3 with the backed up db.sqlite3\n')
            f.write('    3. Restart the Django server\n\n')

            f.write('  To restore from Excel:\n')
            f.write('    1. Use the Excel file to review and manually re-enter data\n')
            f.write('    2. Or use a custom import script\n\n')

        self.stdout.write(self.style.SUCCESS('  ✓ Summary created\n'))
        return summary_path

    def create_zip_archive(self, backup_dir, timestamp):
        """Create a ZIP archive of the backup directory"""
        self.stdout.write('[4/4] Creating ZIP archive...')

        zip_path = f'{backup_dir}.zip'

        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, dirs, files in os.walk(backup_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, os.path.dirname(backup_dir))
                    zipf.write(file_path, arcname)

        zip_size = os.path.getsize(zip_path) / (1024 * 1024)  # Convert to MB
        self.stdout.write(self.style.SUCCESS(f'  ✓ ZIP created: {zip_size:.2f} MB\n'))

        return zip_path

    def send_backup_email(self, recipient_email, attachment_path, timestamp):
        """Send backup via email (when email is configured)"""
        self.stdout.write(f'\n[EMAIL] Sending backup to {recipient_email}...')

        try:
            subject = f'School Management System - Database Backup ({timestamp})'
            message = f'''
Database backup completed successfully.

Backup Date: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
Backup ID: {timestamp}

Please find the backup file attached.

This backup includes:
- All database tables in Excel format
- Raw SQLite database file
- Backup summary

To restore, please refer to the backup_summary.txt file included in the archive.

---
School Management System - Automated Backup
            '''

            email = EmailMessage(
                subject=subject,
                body=message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[recipient_email],
            )

            # Attach the file
            with open(attachment_path, 'rb') as f:
                file_name = os.path.basename(attachment_path)
                email.attach(file_name, f.read())

            email.send()

            self.stdout.write(self.style.SUCCESS(f'  ✓ Email sent successfully to {recipient_email}\n'))

        except Exception as e:
            self.stdout.write(self.style.ERROR(f'  ✗ Email failed: {str(e)}'))
            self.stdout.write(self.style.WARNING('  ℹ To enable email, configure EMAIL settings in settings.py\n'))

    def get_models_to_export(self):
        """Get all models to export, excluding Django internal tables"""
        models_to_export = []

        # Apps to include
        target_apps = ['users', 'academics', 'schooladmin', 'attendance', 'logs']

        for app_config in apps.get_app_configs():
            if app_config.label in target_apps:
                for model in app_config.get_models():
                    models_to_export.append((app_config.label, model))

        return sorted(models_to_export, key=lambda x: (x[0], x[1]._meta.model_name))
